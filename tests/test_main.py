import datetime
import io

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import main
from main import _cell, _infer_type, _map_headers, _norm_header, _split_samples


class TestNormHeader:
    def test_lowercases_and_strips_turkish_accents(self):
        assert _norm_header("Kolon Adı") == "kolonadi"
        assert _norm_header("ŞEMA") == "sema"

    def test_none_becomes_empty(self):
        assert _norm_header(None) == ""


class TestMapHeaders:
    def test_turkish_headers(self):
        headers = ["Sunucu Ad", "Veri Tabanı Ad", "Şema Ad", "Tablo Ad", "Kolon Ad",
                   "Kolon Sıra No", "Kolon Veri Tipi", "Uzunluk", "Kuruş",
                   "Null Flag", "PK Flag"]
        mapping = _map_headers(headers)
        assert mapping[headers.index("Kolon Ad")] == "kolon"
        assert mapping[headers.index("Tablo Ad")] == "tablo"
        assert mapping[headers.index("Şema Ad")] == "sema"

    def test_english_headers(self):
        headers = ["SERVER", "DATABASE", "SCHEMA", "TABLE_NAME", "COLUMN_NAME", "DATATYPE"]
        mapping = _map_headers(headers)
        assert mapping[4] == "kolon"
        assert mapping[3] == "tablo"
        assert mapping[2] == "sema"

    def test_alternate_column_header_spelling(self):
        # README: "Kolon Ad", "Kolon Adı", "COLUMN_NAME" hepsi çalışmalı
        for label in ("Kolon Ad", "Kolon Adı", "COLUMN_NAME", "column_name"):
            mapping = _map_headers([label])
            assert mapping[0] == "kolon"

    def test_missing_column_header_raises(self):
        with pytest.raises(HTTPException) as exc:
            _map_headers(["Sunucu", "Veri Tabanı", "Şema"])
        assert exc.value.status_code == 400

    def test_first_matching_column_wins_when_duplicates(self):
        # Aynı alan için birden fazla aday sütun varsa ilk eşleşen bağlanır
        mapping = _map_headers(["Kolon Ad", "Kolon Adı (2)"])
        assert mapping == {0: "kolon"}

    def test_sira_no_does_not_collide_with_sema(self):
        # "sira" kelimesi "sema"nın alt dizgisi değil ama regresyon için garanti altına al
        mapping = _map_headers(["Şema Ad", "Kolon Sıra No", "Kolon Ad"])
        assert mapping[0] == "sema"
        assert mapping[1] == "sira"
        assert mapping[2] == "kolon"


class TestOrnekDegerlerHeader:
    def test_sample_column_recognized(self):
        mapping = _map_headers(["Kolon Ad", "Örnek Değerler"])
        assert mapping[1] == "ornek_degerler"

    def test_sample_column_english(self):
        mapping = _map_headers(["COLUMN_NAME", "SAMPLE_VALUES"])
        assert mapping[1] == "ornek_degerler"


class TestVeriNHeaders:
    def test_full_new_inventory_format(self):
        # Kullanıcının yeni envanter şablonu: tablo bilgisi + veri1…veri5 aynı dosyada
        headers = ["Sunucu Ad", "Veri Tabanı Ad", "Şema Ad", "Tablo Ad", "Kolon Ad",
                   "Kolon Sıra No", "Kolon Veri Tipi", "Kolon Veri Tipi Uzunluk",
                   "Kolon Veri Tipi Kuruş", "Kolon Null Flag \n(1-Nullable, 0-Not Null)",
                   "Birincil Anahtar\n/Primary Key Flag \n(1-PK, 0-PK değil)",
                   "veri1", "veri2", "veri3", "veri4", "veri5"]
        mapping = _map_headers(headers)
        assert mapping[0] == "sunucu"
        assert mapping[1] == "veritabani"
        assert mapping[2] == "sema"
        assert mapping[3] == "tablo"
        assert mapping[4] == "kolon"
        assert mapping[5] == "sira"
        assert mapping[6] == "veri_tipi"
        assert mapping[7] == "uzunluk"
        assert mapping[8] == "kurus"
        assert mapping[9] == "nullable"
        assert mapping[10] == "pk"
        assert all(mapping[i] == "veri_n" for i in range(11, 16))

    def test_veri_n_not_confused_with_veri_tipi_or_veritabani(self):
        mapping = _map_headers(["Kolon Ad", "Veri Tabanı Ad", "Kolon Veri Tipi", "veri1"])
        assert mapping[1] == "veritabani"
        assert mapping[2] == "veri_tipi"
        assert mapping[3] == "veri_n"

    def test_veri_n_with_space_variant(self):
        mapping = _map_headers(["Kolon Ad", "Veri 1", "VERİ 2"])
        assert mapping[1] == "veri_n"
        assert mapping[2] == "veri_n"


class TestSplitSamples:
    def test_semicolon_and_pipe_separators(self):
        assert _split_samples("a; b | c") == ["a", "b", "c"]

    def test_comma_needs_trailing_space(self):
        # "1,5" gibi ondalıklar bölünmez; ", " ile ayrılmış listeler bölünür
        assert _split_samples("1,5") == ["1,5"]
        assert _split_samples("elma, armut") == ["elma", "armut"]

    def test_empty_and_blank(self):
        assert _split_samples("") == []
        assert _split_samples("  ;  ") == []


NEW_FORMAT_HEADERS = [
    "Sunucu Ad", "Veri Tabanı Ad", "Şema Ad", "Tablo Ad", "Kolon Ad",
    "Kolon Sıra No", "Kolon Veri Tipi", "Kolon Veri Tipi Uzunluk",
    "Kolon Veri Tipi Kuruş", "Kolon Null Flag \n(1-Nullable, 0-Not Null)",
    "Birincil Anahtar\n/Primary Key Flag \n(1-PK, 0-PK değil)",
    "veri1", "veri2", "veri3", "veri4", "veri5",
]


class TestUploadExportNewFormat:
    """Yeni envanter şablonu (tablo bilgisi + veri1…veri5) upload → export tam turu."""

    def _xlsx(self, rows):
        wb = Workbook()
        ws = wb.active
        ws.append(NEW_FORMAT_HEADERS)
        for r in rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_upload_collects_veri_columns_as_samples(self):
        client = TestClient(main.app)
        buf = self._xlsx([
            ["srv1", "coredb", "core", "MusteriHesap", "mhIbanNo", 5, "varchar", 26, "", 1, 0,
             "TR330006100519786457841326", "TR560011100000000012345678", None, None, None],
            ["srv1", "coredb", "core", "MusteriHesap", "mhTutar", 6, "decimal", 18, 2, 1, 0,
             None, None, None, None, None],
        ])
        resp = client.post("/api/upload", files={"file": ("envanter.xlsx", buf)})
        assert resp.status_code == 200
        data = resp.json()
        assert data["count"] == 2
        assert data["with_samples"] == 1
        r0 = data["rows"][0]
        assert r0["kolon"] == "mhIbanNo"
        assert r0["sunucu"] == "srv1"
        assert r0["veritabani"] == "coredb"
        assert r0["sira"] == "5"
        assert r0["veri_tipi"] == "varchar"
        assert r0["uzunluk"] == "26"
        assert r0["nullable"] == "1"
        assert r0["pk"] == "0"
        assert r0["ornek_degerler"] == [
            "TR330006100519786457841326", "TR560011100000000012345678"]
        assert data["rows"][1]["ornek_degerler"] == []
        assert data["rows"][1]["kurus"] == "2"

    def test_export_preserves_input_structure_with_veri_columns(self):
        client = TestClient(main.app)
        row = {
            "sunucu": "srv1", "veritabani": "coredb", "sema": "core",
            "tablo": "MusteriHesap", "kolon": "mhIbanNo", "sira": "5",
            "veri_tipi": "varchar", "uzunluk": "26", "kurus": "", "nullable": "1",
            "pk": "0", "ornek_degerler": ["TR33", "TR56"],
        }
        result = {"kategoriler": [1, 5], "kategori_adlari": ["Kişisel Veri", "Müşteri Sırrı"],
                  "ana_kategori": 5, "ana_kategori_adi": "Müşteri Sırrı", "teknik": False,
                  "acilim": "müşteri hesap IBAN numarası", "guven": 0.93,
                  "gerekce": "IBAN deseni", "kaynak": "llm"}
        resp = client.post("/api/export", json={"items": [{"row": row, "result": result}]})
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        headers = [c.value for c in ws[1]]
        # Girdi yapısı korunur: 11 temel sütun + veri1/veri2, ardından sınıflandırma
        assert headers[:5] == ["Sunucu Ad", "Veri Tabanı Ad", "Şema Ad", "Tablo Ad", "Kolon Ad"]
        assert headers[7] == "Kolon Veri Tipi Uzunluk"
        assert headers[11] == "veri1"
        assert headers[12] == "veri2"
        values = [c.value for c in ws[2]]
        assert values[4] == "mhIbanNo"
        assert values[11] == "TR33"
        assert values[12] == "TR56"
        ana_idx = headers.index("Ana Kategori")
        assert values[ana_idx] == "5. Müşteri Sırrı"


class TestInferType:
    def test_all_ints(self):
        assert _infer_type([1, 2, 3]) == ("int", "")

    def test_mixed_numeric_is_decimal(self):
        assert _infer_type([1, 2.5]) == ("decimal", "")

    def test_dates(self):
        assert _infer_type([datetime.date(2026, 1, 1)]) == ("date", "")

    def test_text_gets_max_length(self):
        assert _infer_type(["ab", "abcd"]) == ("varchar", "4")

    def test_empty_values(self):
        assert _infer_type([None, ""]) == ("", "")


class TestCell:
    def test_none_becomes_empty_string(self):
        assert _cell(None) == ""

    def test_integer_float_loses_decimal(self):
        assert _cell(26.0) == "26"

    def test_non_integer_float_keeps_value(self):
        assert _cell(26.5) == "26.5"

    def test_string_is_stripped(self):
        assert _cell("  varchar  ") == "varchar"

    def test_int_passthrough(self):
        assert _cell(1) == "1"

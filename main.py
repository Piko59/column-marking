"""Kolon İşaretleme — FastAPI backend.

Çalıştırma:  uvicorn main:app --reload --port 8000
Arayüz:      http://localhost:8000
"""

import io
import unicodedata

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

from benchmark import dataset as bench_dataset
from benchmark import jobs as bench_jobs
from benchmark import store as bench_store
from classifier.categories import CATEGORIES
from classifier.pipeline import classify_rows
from classifier.rules import analyze_column

app = FastAPI(title="Kolon İşaretleme")

MAX_ROWS_PER_REQUEST = 200  # frontend bu boyutta parçalar gönderir


# --- Yardımcılar ---------------------------------------------------------------

def _norm_header(h) -> str:
    """Başlığı eşleştirme için normalize eder: küçük harf, aksansız, boşluksuz."""
    s = str(h or "").lower()
    s = s.replace("ı", "i").replace("ş", "s").replace("ç", "c")
    s = s.replace("ğ", "g").replace("ü", "u").replace("ö", "o")
    s = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in s if ch.isalnum())

# alan -> normalize başlıkta aranacak alt dizgiler (öncelik sırasıyla)
HEADER_RULES: list[tuple[str, list[str]]] = [
    ("sunucu", ["sunucu", "server"]),
    ("veritabani", ["veritabani", "database"]),
    ("sema", ["sema", "schema"]),
    ("tablo", ["tabload", "tablo", "table"]),
    ("sira", ["sirano", "sira", "order", "position"]),
    ("uzunluk", ["uzunluk", "length"]),
    ("kurus", ["kurus", "scale", "precision"]),
    ("nullable", ["null"]),
    ("pk", ["birincil", "primary", "pk"]),
    ("veri_tipi", ["veritipi", "datatype", "tip"]),
    ("kolon", ["kolonad", "kolon", "column"]),
]


def _map_headers(headers: list) -> dict[int, str]:
    """Sütun indeksi -> alan adı eşlemesi. Her alan ilk eşleşen sütuna bağlanır."""
    mapping: dict[int, str] = {}
    assigned: set[str] = set()
    for field, needles in HEADER_RULES:
        for idx, h in enumerate(headers):
            if idx in mapping:
                continue
            nh = _norm_header(h)
            if nh and any(n in nh for n in needles):
                mapping[idx] = field
                assigned.add(field)
                break
    if "kolon" not in assigned:
        raise HTTPException(400, "Excel'de 'Kolon Ad' başlıklı bir sütun bulunamadı.")
    return mapping


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# --- Modeller ------------------------------------------------------------------

class RowIn(BaseModel):
    sunucu: str = ""
    veritabani: str = ""
    sema: str = ""
    tablo: str = ""
    kolon: str
    sira: str = ""
    veri_tipi: str = ""
    uzunluk: str = ""
    kurus: str = ""
    nullable: str = ""
    pk: str = ""
    # İsteğe bağlı örnek değerler (maskeli olarak prompta gider, bkz. classifier/rules.mask_sample).
    ornek_degerler: list[str] = Field(default_factory=list)


class ClassifyRequest(BaseModel):
    rows: list[RowIn] = Field(..., max_length=MAX_ROWS_PER_REQUEST)
    use_judge: bool = True
    mode: str = "name_content"  # "name_only" | "content_only" | "name_content"


class ExportRow(BaseModel):
    row: RowIn
    result: dict | None = None


class ExportRequest(BaseModel):
    items: list[ExportRow]


# --- Uçlar ----------------------------------------------------------------------

@app.get("/api/categories")
def get_categories():
    return {"categories": CATEGORIES}


@app.post("/api/upload")
async def upload_excel(file: UploadFile = File(...)):
    """Excel'i ayrıştırır ve normalize satır listesi döndürür (sınıflandırma yapmaz)."""
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Lütfen .xlsx uzantılı bir dosya yükleyin.")
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Dosya okunamadı; geçerli bir Excel dosyası değil.")
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        raise HTTPException(400, "Excel dosyası boş.")
    mapping = _map_headers(headers)

    rows = []
    for raw in rows_iter:
        row = {field: "" for field in RowIn.model_fields}
        for idx, field in mapping.items():
            if idx < len(raw):
                row[field] = _cell(raw[idx])
        if row["kolon"]:
            rows.append(row)
    wb.close()
    if not rows:
        raise HTTPException(400, "Dosyada sınıflandırılacak satır bulunamadı.")
    return {"rows": rows, "count": len(rows)}


@app.post("/api/classify")
async def classify(req: ClassifyRequest):
    """Satır grubunu sınıflandırır. Frontend büyük dosyaları parça parça gönderir."""
    if not req.rows:
        raise HTTPException(400, "Satır listesi boş.")
    if req.mode not in ("name_only", "content_only", "name_content"):
        raise HTTPException(400, "Geçersiz mode.")
    results = await classify_rows(
        [r.model_dump() for r in req.rows], use_judge=req.use_judge, mode=req.mode
    )
    return {"results": results}


@app.post("/api/analyze")
def analyze(row: RowIn):
    """LLM'siz hızlı analiz: önek çözümü + sözlük ipuçları (tekil sorgu ekranı için)."""
    return analyze_column(row.kolon, row.tablo)


@app.post("/api/export")
def export_excel(req: ExportRequest):
    """Sonuçları orijinal kolonlar + sınıflandırma sütunlarıyla .xlsx olarak indirir."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sınıflandırma"
    base_headers = [
        "Sunucu Ad", "Veri Tabanı Ad", "Şema Ad", "Tablo Ad", "Kolon Ad",
        "Kolon Sıra No", "Kolon Veri Tipi", "Uzunluk", "Kuruş", "Null Flag", "PK Flag",
    ]
    cat_headers = [f"{i}. {name}" for i, name in CATEGORIES.items()]
    ws.append(base_headers + cat_headers
              + ["Ana Kategori", "Olası Kategoriler", "Teknik Kolon",
                 "Tahmini Açılım", "Güven", "Gerekçe", "Kaynak"])

    for item in req.items:
        r, res = item.row, item.result or {}
        cats = set(res.get("kategoriler") or [])
        ana = res.get("ana_kategori")
        acilim = ""
        if res.get("kaynak") and res["kaynak"] != "hata":
            acilim = res.get("acilim") or "açılım bulunamadı"
        ws.append(
            [r.sunucu, r.veritabani, r.sema, r.tablo, r.kolon,
             r.sira, r.veri_tipi, r.uzunluk, r.kurus, r.nullable, r.pk]
            + [1 if i in cats else 0 for i in CATEGORIES]
            + [f"{ana}. {res.get('ana_kategori_adi', '')}" if ana else "",
               ", ".join(res.get("kategori_adlari") or []),
               1 if res.get("teknik") else 0,
               acilim,
               res.get("guven", ""), res.get("gerekce", ""), res.get("kaynak", "")]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="kolon_siniflandirma.xlsx"'},
    )


# --- Benchmark uçları -----------------------------------------------------------

class BenchmarkRunRequest(BaseModel):
    modes: list[str] = Field(default_factory=lambda: ["name_only", "content_only", "name_content"])
    use_judge: bool = True


@app.get("/api/benchmark/dataset")
def benchmark_dataset_summary():
    """Golden veri setinin özeti (kavram/satır sayısı, kovalar) — koşmadan önce UI'da gösterilir."""
    return bench_dataset.dataset_summary()


@app.get("/api/benchmark/dataset/concepts")
def benchmark_dataset_concepts():
    """Veri setindeki her kavramın ground truth'u ve örnek değerleri — şeffaflık için."""
    return {
        "concepts": [
            {
                "key": c.key, "bucket": c.bucket, "ana_kategori": c.ana_kategori,
                "kategoriler": c.kategoriler, "teknik": c.teknik, "gerekce": c.gerekce,
                "named": {"sema": c.named_schema, "tablo": c.named_table, "kolon": c.named_column},
                "random": {"sema": c.random_schema, "tablo": c.random_table, "kolon": c.random_column},
                "veri_tipi": c.veri_tipi, "uzunluk": c.uzunluk, "ornek_degerler": c.sample_values,
            }
            for c in bench_dataset.CONCEPTS
        ]
    }


@app.post("/api/benchmark/run")
async def benchmark_run(req: BenchmarkRunRequest):
    """Benchmark koşusunu arka planda başlatır; ilerleme /api/benchmark/jobs/{id} ile izlenir.

    async def OLMALI: start_job() içi asyncio.create_task() çağırıyor, bu da çalışan bir
    event loop gerektirir. Sync (def) uç noktalar FastAPI'de thread pool'da çalışır ve o
    thread'de event loop yoktur — sync bırakılırsa "no running event loop" ile patlar.
    """
    valid = {"name_only", "content_only", "name_content"}
    if not req.modes or not set(req.modes).issubset(valid):
        raise HTTPException(400, f"Geçersiz mode listesi; beklenen alt küme: {sorted(valid)}")
    job_id = bench_jobs.start_job(req.modes, req.use_judge)
    return {"job_id": job_id}


@app.get("/api/benchmark/jobs/{job_id}")
def benchmark_job_status(job_id: str):
    job = bench_jobs.get_job(job_id)
    if job is None:
        raise HTTPException(404, "İş bulunamadı.")
    return job


@app.get("/api/benchmark/runs")
def benchmark_list_runs(limit: int = 50):
    return {"runs": bench_store.list_runs(limit=limit)}


@app.get("/api/benchmark/runs/{run_id}")
def benchmark_get_run(run_id: str):
    run = bench_store.get_run(run_id)
    if run is None:
        raise HTTPException(404, "Koşu bulunamadı.")
    return run


@app.delete("/api/benchmark/runs/{run_id}")
def benchmark_delete_run(run_id: str):
    if not bench_store.delete_run(run_id):
        raise HTTPException(404, "Koşu bulunamadı.")
    return {"ok": True}


app.mount("/", StaticFiles(directory="static", html=True), name="static")
